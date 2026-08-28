#!/usr/bin/env python3
"""
Kie 账单成本校准工具（月度对账用）。

用法：
    python3 scripts/kie_cost_audit.py <kie账单.xlsx> [更多.xlsx ...]

账单来源：Kie 控制台 → Credits 页 → 导出 xlsx（含 Model/Credits/TaskID 列）。

输出：
  1. 按模型×扣费分组的桶表（次数、credits 合计、折算美元）；
  2. 每个模型折算的 $/秒（有 duration 列时直接算；否则按下方 KNOWN_TIERS 推断）；
  3. 与 lib/video-pricing.ts 成本依据（MODEL_COST_BASIS_USD_PER_SECOND）的对照与建议。

口径：Kie credit = $0.005（官方与 2026-08-27 充值实测一致）；失败任务也计费，
账单里的失败行同样代表真实成本。校准后把结果回填 video-pricing.ts 的
MODEL_COST_BASIS_USD_PER_SECOND（verified 改 true），并跑 video-pricing.test.ts。
"""
import sys
import glob
from collections import defaultdict

CREDIT_USD = 0.005
# 已知计费档（模型 → credits→秒数）。8s 档等新档位发现后补到这里。
KNOWN_TIERS = {
    "veo-3-1": {30: 4},  # 2026-08-27 实测：4s 档 30cr；8s 档待实测（若同为 30cr 则成本减半）
}
# Kie 账单模型名 → lib/video-pricing.ts 的 model key
MODEL_NAME_MAP = {
    "veo-3-1": "veo31(按档位区分 Lite/Fast/Quality，需人工判读)",
    "nano-banana-2": "nanoBanana2",
    "seedance-2-5": "seedance25",
    "kling-3-0": "kling3",
    "wan-2-7": "wan27",
    "minimax-h3": "minimaxH3",
}
# lib/video-pricing.ts 当前成本依据（$/秒）——脚本只读对照，唯一事实源在 TS 里
CURRENT_BASIS = {
    "veo31Lite": 0.0375, "veo31Fast": 0.05, "veo31Quality": 0.25,
    "seedance25": 0.23, "seedance2Fast": 0.05, "seedance2Mini": 0.04,
    "seedance2": 0.09, "kling3": 0.06, "happyHorse": 0.06,
    "wan27": 0.05, "geminiOmni": 0.06, "minimaxH3": 0.07,
}
CURRENT_UNIT = {
    "veo31Lite": 1.5, "veo31Fast": 2, "veo31Quality": 9, "seedance25": 9,
    "seedance2Fast": 2, "seedance2Mini": 1.5, "seedance2": 3.5, "kling3": 2.5,
    "happyHorse": 2.5, "wan27": 2, "geminiOmni": 2.5, "minimaxH3": 2.5,
}


def pick_column(headers, *keywords):
    lowered = [str(h).lower() if h is not None else "" for h in headers]
    for kw in keywords:
        for i, h in enumerate(lowered):
            if kw in h:
                return i
    return None


def load_rows(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = rows[0]
    col_model = pick_column(headers, "model")
    col_credit = pick_column(headers, "credit", "consume", "amount")
    col_status = pick_column(headers, "status", "state", "success", "flag")
    col_duration = pick_column(headers, "duration", "seconds")
    if col_model is None or col_credit is None:
        raise SystemExit(f"[{path}] 找不到 model/credits 列。表头：{headers}")
    out = []
    for r in rows[1:]:
        if r[col_model] is None:
            continue
        out.append({
            "model": str(r[col_model]).strip(),
            "credits": float(r[col_credit] or 0),
            "status": str(r[col_status]).strip().lower() if col_status is not None else "",
            "duration": float(r[col_duration]) if col_duration is not None and r[col_duration] else None,
        })
    return out


def main(paths):
    all_rows = []
    for pattern in paths:
        for p in glob.glob(pattern):
            rows = load_rows(p)
            print(f"已读取 {p}: {len(rows)} 行")
            all_rows.extend(rows)
    if not all_rows:
        raise SystemExit("没有读到任何账单行")

    # 桶：模型 × credits
    buckets = defaultdict(lambda: {"n": 0, "credits": 0.0})
    # 干净速率：按行折算 $/s 再平均（失败损耗由底线公式的 1.7× 承担，不在此混入）
    per_model_rates = defaultdict(list)
    for r in all_rows:
        b = buckets[(r["model"], r["credits"])]
        b["n"] += 1
        b["credits"] += r["credits"]
        if r["duration"]:
            per_model_rates[r["model"]].append((r["credits"] * CREDIT_USD / r["duration"], r["credits"], r["duration"]))

    print("\n== 扣费桶（模型 × 单次credits）==")
    print(f"{'模型':<22}{'单次credits':>10}{'次数':>8}{'credits合计':>12}{'金额$':>10}")
    total_usd = 0.0
    for (model, credits), b in sorted(buckets.items()):
        usd = b["credits"] * CREDIT_USD
        total_usd += usd
        print(f"{model:<22}{credits:>10g}{b['n']:>8}{b['credits']:>12g}{usd:>10.3f}")

    print(f"\n账单总支出: ${total_usd:.2f}")

    print("\n== 干净 $/秒（按行折算，失败损耗不计入）==")
    for m, rates in sorted(per_model_rates.items()):
        avg = sum(x[0] for x in rates) / len(rates)
        example = rates[0]
        print(f"  {m}: {avg:.4f}/s（如 {example[1]:g}cr/{example[2]:g}s，{len(rates)} 行）")

    print("\n== 与 lib/video-pricing.ts 对照 ==")
    print(f"{'模型key':<40}{'当前依据$/s':>12}{'实测$/s':>12}  建议")
    for bill_model, key_hint in sorted(MODEL_NAME_MAP.items()):
        current_keys = [key_hint] if not key_hint.startswith("veo31") else ["veo31Lite", "veo31Fast", "veo31Quality"]
        current = " / ".join(str(CURRENT_BASIS.get(k, "—")) for k in current_keys)
        rate = None
        src = ""
        if per_model_rates.get(bill_model):
            rates = per_model_rates[bill_model]
            rate = sum(x[0] for x in rates) / len(rates)
            src = f"{rates[0][1]:g}cr/{rates[0][2]:g}s"
        else:
            # 无 duration 列时用已知档推断（取该模型任一桶匹配 KNOWN_TIERS）
            for (m, credits), b in buckets.items():
                if m == bill_model and credits in KNOWN_TIERS.get(bill_model, {}):
                    rate = credits * CREDIT_USD / KNOWN_TIERS[bill_model][credits]
                    src = f"{credits:g}cr/{KNOWN_TIERS[bill_model][credits]:g}s 已知档"
                    break
        if rate is None:
            print(f"{key_hint:<40}{current:>12}{'—':>12}  账单无 duration/已知档，人工判读桶表")
            continue
        unit = CURRENT_UNIT.get(key_hint)
        if unit:
            floor = rate * 1.7 * 2 * 1.04 / 0.10  # 与 minUnitPoints 相同的底线公式
            advice = f"实测 ${rate:.4f}/s（{src}）；底线单价 ≥ {floor:.1f} 分/秒，当前 {unit}"
        else:
            advice = f"实测 ${rate:.4f}/s（{src}）"
        print(f"{key_hint:<40}{current:>12}{rate:>12.4f}  {advice}")

    print("\n下一步：把实测值回填 lib/video-pricing.ts 的 MODEL_COST_BASIS_USD_PER_SECOND（verified: true），跑 video-pricing.test.ts。")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    main(sys.argv[1:])
